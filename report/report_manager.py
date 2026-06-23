# report/report_manager.py
import os
import re
import zipfile
from collections import Counter
from datetime import datetime
from html import escape as html_escape
from xml.sax.saxutils import escape as xml_escape

from models.data_models import ScanSummary


class ReportManager:
    RISK_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "": 4}
    HIGH_RISK_MARKERS = (
        "隐藏",
        "后缀",
        "加密",
        "无法读取",
        "无法访问",
        "OCR失败",
        "任务失败",
        "解析失败",
    )

    def __init__(self, output_dir="audit_reports"):
        """初始化报告管理器，确保输出目录存在于项目根目录"""
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generate_excel_report(self, summary: ScanSummary) -> str:
        """接收 ScanSummary，生成 summary/findings/errors 三个 Sheet 的 Excel 报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_task_name = self._safe_filename(summary.task_name)
        file_name = f"DLP_Audit_{safe_task_name}_{timestamp}.xlsx"
        file_path = os.path.join(self.output_dir, file_name)

        sheets = self._build_report_sheets(summary)

        try:
            self._write_with_pandas(file_path, sheets)
            return os.path.abspath(file_path)
        except ImportError:
            self._write_minimal_xlsx(file_path, sheets)
            return os.path.abspath(file_path)
        except Exception as e:
            print(f"\n生成报告时发生错误: {e}")
            return ""

    def generate_combined_excel_report(self, summaries: list[ScanSummary]) -> str:
        """生成综合检查报告：overall_summary + all_findings + all_errors + 各任务 summary。"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"DLP_Audit_综合涉密信息检查_{timestamp}.xlsx"
        file_path = os.path.join(self.output_dir, file_name)

        sheets = self._build_combined_report_sheets(summaries)
        try:
            self._write_with_pandas(file_path, sheets)
            return os.path.abspath(file_path)
        except ImportError:
            self._write_minimal_xlsx(file_path, sheets)
            return os.path.abspath(file_path)
        except Exception as e:
            print(f"\n生成综合报告时发生错误: {e}")
            return ""

    def generate_html_report(self, summary: ScanSummary, excel_path: str | None = None) -> str:
        """生成单任务 HTML 摘要报告。Excel 仍作为完整明细归档。"""
        html_path = self._html_path_for(summary.task_name, excel_path)
        try:
            html = self._build_html_report(summary.task_name, [summary], excel_path)
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            return os.path.abspath(html_path)
        except Exception as e:
            print(f"\n生成 HTML 摘要报告时发生错误: {e}")
            return ""

    def generate_combined_html_report(self, summaries: list[ScanSummary],
                                      excel_path: str | None = None) -> str:
        """生成综合任务 HTML 摘要报告。"""
        html_path = self._html_path_for("综合涉密信息检查", excel_path)
        try:
            html = self._build_html_report("综合涉密信息检查", summaries, excel_path)
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            return os.path.abspath(html_path)
        except Exception as e:
            print(f"\n生成综合 HTML 摘要报告时发生错误: {e}")
            return ""

    def _build_combined_report_sheets(self, summaries: list[ScanSummary]) -> dict[str, list[list[object]]]:
        total_scanned = sum(summary.total_scanned for summary in summaries)
        all_results = [res for summary in summaries for res in summary.results]
        findings = self._sort_results([res for res in all_results if not res.error_msg])
        errors = [res for res in all_results if res.error_msg]

        overall_rows = [["metric", "value"]]
        overall_rows.extend([
            ["task_name", "综合涉密信息检查"],
            ["task_count", len(summaries)],
            ["total_scanned", total_scanned],
            ["total_findings", len(findings)],
            ["total_errors", len(errors)],
        ])
        for summary in summaries:
            overall_rows.append([f"{summary.task_name} - scanned", summary.total_scanned])
            overall_rows.append([f"{summary.task_name} - findings", len([r for r in summary.results if not r.error_msg])])
            overall_rows.append([f"{summary.task_name} - errors", len([r for r in summary.results if r.error_msg])])
            if summary.scanned_details:
                details = "\n".join([f"{key}: {val}" for key, val in summary.scanned_details.items()])
                overall_rows.append([f"{summary.task_name} - details", details])

        finding_rows = [[
            "task_name",
            "source_type",
            "source_path",
            "location",
            "rule_id",
            "rule_name",
            "risk_level",
            "keyword",
            "context",
            "rule_description",
        ]]
        error_rows = [[
            "task_name",
            "source_type",
            "source_path",
            "location",
            "keyword",
            "error_msg",
            "context",
        ]]

        task_name_by_result = {
            id(res): summary.task_name
            for summary in summaries
            for res in summary.results
        }

        for res in findings:
            finding_rows.append([
                task_name_by_result.get(id(res), ""),
                res.source_type,
                res.source_path,
                res.line_number,
                res.rule_id,
                res.rule_name,
                res.risk_level,
                res.keyword,
                res.context,
                res.rule_description,
            ])

        for summary in summaries:
            for res in summary.results:
                if res.error_msg:
                    error_rows.append([
                        summary.task_name,
                        res.source_type,
                        res.source_path,
                        res.line_number,
                        res.keyword,
                        res.error_msg,
                        res.context,
                    ])

        sheets = {
            "overall_summary": overall_rows,
            "high_risk_findings": self._build_high_risk_rows(summaries),
            "all_findings": finding_rows,
            "all_errors": error_rows,
        }
        sheets.update(self._build_metadata_sheets(summaries))

        # 保留每个子任务的 summary，便于老师快速核对每个任务的扫描数量
        for index, summary in enumerate(summaries, start=1):
            sheet_name = self._safe_sheet_name(f"{index}_{summary.task_name}_summary")
            sheets[sheet_name] = self._build_report_sheets(summary)["summary"]

        return sheets

    def _build_report_sheets(self, summary: ScanSummary) -> dict[str, list[list[object]]]:
        findings = self._sort_results([res for res in summary.results if not res.error_msg])
        errors = [res for res in summary.results if res.error_msg]

        details_list = [f"{key}: {val}" for key, val in summary.scanned_details.items()]
        details_str = "\n".join(details_list) if details_list else "无详细分类"

        summary_rows = [
            ["metric", "value"],
            ["task_name", summary.task_name],
            ["total_scanned", summary.total_scanned],
            ["total_findings", len(findings)],
            ["total_errors", len(errors)],
            ["scanned_details", details_str],
        ]

        finding_rows = [[
            "source_type",
            "source_path",
            "location",
            "rule_id",
            "rule_name",
            "risk_level",
            "keyword",
            "context",
            "rule_description",
        ]]
        for res in findings:
            finding_rows.append([
                res.source_type,
                res.source_path,
                res.line_number,
                res.rule_id,
                res.rule_name,
                res.risk_level,
                res.keyword,
                res.context,
                res.rule_description,
            ])

        error_rows = [[
            "source_type",
            "source_path",
            "location",
            "keyword",
            "error_msg",
            "context",
        ]]
        for res in errors:
            error_rows.append([
                res.source_type,
                res.source_path,
                res.line_number,
                res.keyword,
                res.error_msg,
                res.context,
            ])

        sheets = {
            "summary": summary_rows,
            "high_risk_findings": self._build_high_risk_rows([summary]),
            "findings": finding_rows,
            "errors": error_rows,
        }
        sheets.update(self._build_metadata_sheets([summary]))
        return sheets

    def _build_metadata_sheets(self, summaries: list[ScanSummary]) -> dict[str, list[list[object]]]:
        sheets: dict[str, list[list[object]]] = {}

        web_rows = [[
            "task_name",
            "url",
            "depth",
            "status_code",
            "content_type",
            "title",
            "text_sha256",
            "text_chars",
            "fetched_at",
            "error_msg",
        ]]
        db_rows = [[
            "task_name",
            "label",
            "host",
            "port",
            "database",
            "user",
            "status",
            "total_scanned",
            "total_findings",
            "total_errors",
            "error_msg",
        ]]

        for summary in summaries:
            for snapshot in summary.metadata.get("web_snapshots", []):
                web_rows.append([
                    summary.task_name,
                    snapshot.get("url", ""),
                    snapshot.get("depth", ""),
                    snapshot.get("status_code", ""),
                    snapshot.get("content_type", ""),
                    snapshot.get("title", ""),
                    snapshot.get("text_sha256", ""),
                    snapshot.get("text_chars", ""),
                    snapshot.get("fetched_at", ""),
                    snapshot.get("error_msg", ""),
                ])
            for target in summary.metadata.get("db_targets", []):
                db_rows.append([
                    summary.task_name,
                    target.get("label", ""),
                    target.get("host", ""),
                    target.get("port", ""),
                    target.get("database", ""),
                    target.get("user", ""),
                    target.get("status", ""),
                    target.get("total_scanned", ""),
                    target.get("total_findings", ""),
                    target.get("total_errors", ""),
                    target.get("error_msg", ""),
                ])

        if len(web_rows) > 1:
            sheets["web_snapshots"] = web_rows
        if len(db_rows) > 1:
            sheets["db_targets"] = db_rows
        return sheets

    def _write_with_pandas(self, file_path: str, sheets: dict[str, list[list[object]]]) -> None:
        import pandas as pd
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, rows in sheets.items():
                header = rows[0]
                body = rows[1:]
                df = pd.DataFrame(body, columns=header)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                self._format_openpyxl_sheet(
                    worksheet,
                    rows,
                    header,
                    Font,
                    PatternFill,
                    Alignment,
                    FormulaRule,
                )

    def _write_minimal_xlsx(self, file_path: str, sheets: dict[str, list[list[object]]]) -> None:
        """Write a small standards-compliant XLSX without pandas/openpyxl."""
        with zipfile.ZipFile(file_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", self._content_types_xml(len(sheets)))
            archive.writestr("_rels/.rels", self._root_rels_xml())
            archive.writestr("xl/workbook.xml", self._workbook_xml(list(sheets.keys())))
            archive.writestr("xl/_rels/workbook.xml.rels", self._workbook_rels_xml(len(sheets)))

            for index, rows in enumerate(sheets.values(), start=1):
                archive.writestr(f"xl/worksheets/sheet{index}.xml", self._worksheet_xml(rows))

    def _worksheet_xml(self, rows: list[list[object]]) -> str:
        row_xml = []
        for row_idx, row in enumerate(rows, start=1):
            cells = []
            for col_idx, value in enumerate(row, start=1):
                cell_ref = f"{self._column_name(col_idx)}{row_idx}"
                cell_value = self._clean_cell(value)
                cells.append(
                    f'<c r="{cell_ref}" t="inlineStr"><is><t>{xml_escape(cell_value)}</t></is></c>'
                )
            row_xml.append(f'<row r="{row_idx}">{"".join(cells)}</row>')

        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<sheetData>{"".join(row_xml)}</sheetData>'
            '</worksheet>'
        )

    def _content_types_xml(self, sheet_count: int) -> str:
        overrides = [
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
            '<Default Extension="xml" ContentType="application/xml"/>',
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
        ]
        for index in range(1, sheet_count + 1):
            overrides.append(
                f'<Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            f'{"".join(overrides)}'
            '</Types>'
        )

    def _root_rels_xml(self) -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>'
        )

    def _workbook_xml(self, sheet_names: list[str]) -> str:
        sheet_xml = []
        for index, sheet_name in enumerate(sheet_names, start=1):
            sheet_xml.append(
                f'<sheet name="{xml_escape(sheet_name)}" sheetId="{index}" r:id="rId{index}"/>'
            )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets>{"".join(sheet_xml)}</sheets>'
            '</workbook>'
        )

    def _workbook_rels_xml(self, sheet_count: int) -> str:
        rels = []
        for index in range(1, sheet_count + 1):
            rels.append(
                f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
            )
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'{"".join(rels)}'
            '</Relationships>'
        )

    def _column_name(self, index: int) -> str:
        name = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name
        return name

    def _clean_cell(self, value: object) -> str:
        text = "" if value is None else str(value)
        return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', text)

    def _safe_sheet_name(self, value: str) -> str:
        cleaned = re.sub(r'[\\/:*?"<>|\[\]]+', '_', value).strip()
        return (cleaned or "sheet")[:31]

    def _safe_filename(self, value: str) -> str:
        cleaned = re.sub(r'[\\/:*?"<>|]+', '_', value)
        return cleaned.strip() or "DLP_Audit"

    def _sort_results(self, results):
        return sorted(
            results,
            key=lambda res: (
                self._risk_rank(res),
                str(res.source_type),
                str(res.source_path),
                str(res.line_number),
                str(res.keyword),
            ),
        )

    def _risk_rank(self, result) -> int:
        if result.error_msg:
            return 1
        return self.RISK_ORDER.get(str(result.risk_level).lower(), 4)

    def _is_high_risk_result(self, result) -> bool:
        risk_level = str(result.risk_level).lower()
        text = " ".join([
            str(result.keyword or ""),
            str(result.rule_id or ""),
            str(result.rule_name or ""),
            str(result.error_msg or ""),
        ])
        return (
            bool(result.error_msg)
            or risk_level in {"critical", "high"}
            or str(result.rule_id or "").startswith("SYSTEM_")
            or any(marker in text for marker in self.HIGH_RISK_MARKERS)
        )

    def _build_high_risk_rows(self, summaries: list[ScanSummary]) -> list[list[object]]:
        rows = [[
            "task_name",
            "source_type",
            "source_path",
            "location",
            "risk_level",
            "rule_id",
            "rule_name",
            "keyword",
            "context",
            "error_msg",
            "rule_description",
        ]]
        high_risk_items = []
        for summary in summaries:
            for res in summary.results:
                if self._is_high_risk_result(res):
                    high_risk_items.append((summary.task_name, res))

        for task_name, res in sorted(high_risk_items, key=lambda item: self._risk_rank(item[1])):
            rows.append([
                task_name,
                res.source_type,
                res.source_path,
                res.line_number,
                res.risk_level or ("error" if res.error_msg else ""),
                res.rule_id or ("SYSTEM_PARSE_OR_ACCESS_ERROR" if res.error_msg else ""),
                res.rule_name or ("解析/访问异常" if res.error_msg else ""),
                res.keyword,
                res.context,
                res.error_msg,
                res.rule_description,
            ])
        return rows

    def _html_path_for(self, task_name: str, excel_path: str | None = None) -> str:
        if excel_path:
            return os.path.splitext(excel_path)[0] + ".html"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_task_name = self._safe_filename(task_name)
        return os.path.join(self.output_dir, f"DLP_Audit_{safe_task_name}_{timestamp}.html")

    def _build_html_report(self, title: str, summaries: list[ScanSummary],
                           excel_path: str | None = None) -> str:
        all_results = [res for summary in summaries for res in summary.results]
        findings = self._sort_results([res for res in all_results if not res.error_msg])
        errors = [res for res in all_results if res.error_msg]
        high_risk_items = [
            (summary.task_name, res)
            for summary in summaries
            for res in summary.results
            if self._is_high_risk_result(res)
        ]
        error_items = [
            (summary.task_name, res)
            for summary in summaries
            for res in summary.results
            if res.error_msg
        ]
        risk_counts = Counter((res.risk_level or "unknown").lower() for res in findings)
        rule_counts = Counter(
            (res.rule_id or "-", res.rule_name or res.keyword or "-")
            for res in findings
        )

        task_rows = []
        for summary in summaries:
            summary_findings = [res for res in summary.results if not res.error_msg]
            summary_errors = [res for res in summary.results if res.error_msg]
            task_rows.append([
                summary.task_name,
                summary.total_scanned,
                len(summary_findings),
                len(summary_errors),
                self._summarize_details_for_html(summary.scanned_details),
            ])

        high_risk_rows = [
            self._html_result_row(task_name, res)
            for task_name, res in sorted(
                high_risk_items,
                key=lambda item: self._risk_rank(item[1])
            )[:100]
        ]
        error_rows = [
            [
                task_name,
                res.source_type,
                res.source_path,
                res.line_number,
                res.keyword,
                res.error_msg,
            ]
            for task_name, res in error_items[:100]
        ]
        top_rule_rows = [
            [rule_id, rule_name, count]
            for (rule_id, rule_name), count in rule_counts.most_common(20)
        ]
        risk_rows = [
            [risk, risk_counts.get(risk, 0)]
            for risk in ["critical", "high", "medium", "low", "unknown"]
            if risk_counts.get(risk, 0)
        ]
        web_snapshot_rows = self._web_snapshot_rows_for_html(summaries)
        db_target_rows = self._db_target_rows_for_html(summaries)

        excel_note = (
            f"完整明细见同目录 Excel 文件：{html_escape(os.path.basename(excel_path))}"
            if excel_path else
            "完整明细见同目录 Excel 文件。"
        )
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        report_title = f"{title}摘要报告"

        return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>{html_escape(report_title)}</title>
  <style>
    body {{ margin: 0; font-family: "Microsoft YaHei", Arial, sans-serif; background: #f5f7fb; color: #1f2937; }}
    .page {{ max-width: 1180px; margin: 0 auto; padding: 28px; }}
    .hero {{ background: #0f172a; color: #ffffff; padding: 24px 28px; border-radius: 8px; }}
    .hero h1 {{ margin: 0 0 8px; font-size: 26px; }}
    .hero p {{ margin: 4px 0; color: #dbeafe; }}
    .cards {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin: 18px 0; }}
    .toolbar {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; background: #ffffff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px; margin: 16px 0; position: sticky; top: 0; z-index: 5; }}
    .toolbar input, .toolbar select {{ border: 1px solid #cbd5e1; border-radius: 6px; padding: 8px 10px; font-size: 14px; }}
    .toolbar input {{ min-width: 260px; flex: 1; }}
    .toolbar a {{ color: #2563eb; text-decoration: none; font-size: 13px; }}
    .quick-nav {{ display: flex; flex-wrap: wrap; gap: 8px; }}
    .quick-nav a {{ background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 6px; padding: 6px 8px; }}
    .card {{ background: #ffffff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 16px; }}
    .card .label {{ color: #64748b; font-size: 13px; }}
    .card .value {{ font-size: 26px; font-weight: 700; margin-top: 6px; }}
    section {{ background: #ffffff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 18px; margin: 16px 0; }}
    h2 {{ font-size: 18px; margin: 0 0 12px; }}
    .table-wrap {{ overflow-x: auto; }}
    table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
    th, td {{ border-bottom: 1px solid #e5e7eb; padding: 8px 10px; text-align: left; vertical-align: top; word-break: break-word; }}
    th {{ background: #f1f5f9; color: #334155; font-weight: 700; }}
    tr.critical td {{ background: #fff1f2; }}
    tr.high td {{ background: #fff7ed; }}
    tr.medium td {{ background: #fefce8; }}
    .muted {{ color: #64748b; }}
    .note {{ font-size: 13px; color: #475569; line-height: 1.7; }}
    .hidden-row {{ display: none; }}
    @media (max-width: 820px) {{ .cards {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }} .page {{ padding: 16px; }} }}
  </style>
</head>
<body>
  <div class="page">
    <div class="hero">
      <h1>{html_escape(report_title)}</h1>
      <p>生成时间：{html_escape(generated_at)}</p>
    </div>
    <div class="cards">
      {self._html_card("扫描任务数", len(summaries))}
      {self._html_card("扫描对象总数", sum(summary.total_scanned for summary in summaries))}
      {self._html_card("命中总数", len(findings))}
      {self._html_card("异常总数", len(errors))}
    </div>
    <div class="toolbar">
      <input id="reportSearch" type="search" placeholder="搜索路径、规则、上下文或异常信息">
      <select id="riskFilter" aria-label="风险等级筛选">
        <option value="">全部风险</option>
        <option value="critical">critical</option>
        <option value="high">high</option>
        <option value="medium">medium</option>
        <option value="low">low</option>
        <option value="error">error</option>
      </select>
      <div class="quick-nav">
        <a href="#module-stats">分模块</a>
        <a href="#risk-dist">风险分布</a>
        <a href="#rule-top">规则 Top</a>
        <a href="#high-risk">高风险</a>
        <a href="#errors">异常</a>
        <a href="#snapshots">快照/DB</a>
      </div>
    </div>
    <section id="module-stats">
      <h2>分模块统计</h2>
      {self._html_table(["任务", "扫描对象", "命中", "异常", "细分统计"], task_rows)}
    </section>
    <section id="risk-dist">
      <h2>风险等级分布</h2>
      {self._html_table(["风险等级", "命中数"], risk_rows)}
    </section>
    <section id="rule-top">
      <h2>规则命中 Top 20</h2>
      {self._html_table(["规则 ID", "规则名称", "命中数"], top_rule_rows)}
    </section>
    <section id="high-risk">
      <h2>高风险明细（前 100 条）</h2>
      {self._html_result_table(high_risk_rows)}
    </section>
    <section id="errors">
      <h2>异常摘要（前 100 条）</h2>
      {self._html_table(["任务", "来源类型", "路径/URL/对象", "位置", "特征", "异常信息"], error_rows, risk_col=6)}
    </section>
    <section id="snapshots">
      <h2>页面快照与数据库目标</h2>
      <h3>Web 页面快照</h3>
      {self._html_table(["任务", "URL", "深度", "状态码", "标题", "文本 SHA-256", "文本长度", "抓取时间", "异常"], web_snapshot_rows, risk_col=8)}
      <h3>数据库目标</h3>
      {self._html_table(["任务", "标签", "Host", "Port", "Database", "User", "状态", "扫描表数", "命中数", "异常数", "异常信息"], db_target_rows, risk_col=6)}
    </section>
    <section>
      <h2>报告说明</h2>
      <p class="note">{excel_note}</p>
      <p class="note">本页面仅展示主要统计、规则分布和前 100 条重点记录，便于快速阅读；全部记录和完整字段以 Excel 明细报告为准。</p>
    </section>
  </div>
  <script>
    (function() {{
      const searchInput = document.getElementById('reportSearch');
      const riskFilter = document.getElementById('riskFilter');
      function applyFilters() {{
        const query = (searchInput.value || '').toLowerCase().trim();
        const risk = (riskFilter.value || '').toLowerCase();
        document.querySelectorAll('tbody tr').forEach(function(row) {{
          const haystack = (row.dataset.search || row.innerText || '').toLowerCase();
          const rowRisk = (row.dataset.risk || '').toLowerCase();
          const matchesText = !query || haystack.indexOf(query) !== -1;
          const matchesRisk = !risk || rowRisk === risk;
          row.classList.toggle('hidden-row', !(matchesText && matchesRisk));
        }});
      }}
      searchInput.addEventListener('input', applyFilters);
      riskFilter.addEventListener('change', applyFilters);
    }})();
  </script>
</body>
</html>
"""

    def _web_snapshot_rows_for_html(self, summaries: list[ScanSummary]) -> list[list[object]]:
        rows: list[list[object]] = []
        for summary in summaries:
            for snapshot in summary.metadata.get("web_snapshots", []):
                rows.append([
                    summary.task_name,
                    snapshot.get("url", ""),
                    snapshot.get("depth", ""),
                    snapshot.get("status_code", ""),
                    snapshot.get("title", ""),
                    snapshot.get("text_sha256", ""),
                    snapshot.get("text_chars", ""),
                    snapshot.get("fetched_at", ""),
                    "error" if snapshot.get("error_msg") else "",
                ])
        return rows[:100]

    def _db_target_rows_for_html(self, summaries: list[ScanSummary]) -> list[list[object]]:
        rows: list[list[object]] = []
        for summary in summaries:
            for target in summary.metadata.get("db_targets", []):
                rows.append([
                    summary.task_name,
                    target.get("label", ""),
                    target.get("host", ""),
                    target.get("port", ""),
                    target.get("database", ""),
                    target.get("user", ""),
                    target.get("status", ""),
                    target.get("total_scanned", ""),
                    target.get("total_findings", ""),
                    target.get("total_errors", ""),
                    target.get("error_msg", ""),
                ])
        return rows[:100]

    def _html_card(self, label: str, value: object) -> str:
        return (
            '<div class="card">'
            f'<div class="label">{html_escape(str(label))}</div>'
            f'<div class="value">{html_escape(str(value))}</div>'
            '</div>'
        )

    def _summarize_details_for_html(self, details: dict[str, int], limit: int = 8) -> str:
        if not details:
            return "-"
        items = list(details.items())
        shown = "; ".join(f"{key}: {value}" for key, value in items[:limit])
        if len(items) > limit:
            shown += f"; 其余 {len(items) - limit} 项见 Excel 明细"
        return shown

    def _html_result_row(self, task_name: str, res) -> list[object]:
        return [
            task_name,
            res.source_type,
            res.source_path,
            res.line_number,
            res.risk_level or ("error" if res.error_msg else ""),
            res.rule_id or ("SYSTEM_PARSE_OR_ACCESS_ERROR" if res.error_msg else ""),
            res.rule_name or ("解析/访问异常" if res.error_msg else ""),
            res.keyword,
            res.context,
            res.error_msg,
        ]

    def _html_result_table(self, rows: list[list[object]]) -> str:
        headers = ["任务", "来源类型", "路径/URL/对象", "位置", "风险等级", "规则 ID", "规则名称", "命中特征", "上下文", "异常信息"]
        return self._html_table(headers, rows, risk_col=4)

    def _html_table(self, headers: list[str], rows: list[list[object]], risk_col: int | None = None) -> str:
        if not rows:
            return '<p class="muted">无记录</p>'
        header_html = "".join(f"<th>{html_escape(str(header))}</th>" for header in headers)
        body_html = []
        for row in rows:
            risk_value = ""
            row_class = ""
            if risk_col is not None and len(row) > risk_col:
                risk_value = str(row[risk_col]).lower()
                if risk_value in {"critical", "high", "medium"}:
                    row_class = f' class="{risk_value}"'
                elif risk_value in {"error", "异常", "failed"}:
                    risk_value = "error"
            elif risk_col is not None and risk_col >= len(row):
                risk_value = "error"
            row_text = " ".join(str(cell) for cell in row)
            cells = "".join(f"<td>{html_escape(str(cell))}</td>" for cell in row)
            body_html.append(
                f'<tr{row_class} data-risk="{html_escape(risk_value)}" '
                f'data-search="{html_escape(row_text)}">{cells}</tr>'
            )
        return (
            f'<div class="table-wrap"><table class="data-table">'
            f'<thead><tr>{header_html}</tr></thead><tbody>{"".join(body_html)}</tbody>'
            f'</table></div>'
        )

    def _format_openpyxl_sheet(self, worksheet, rows, header, Font, PatternFill, Alignment, FormulaRule) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2937")
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        risk_fills = {
            "critical": PatternFill("solid", fgColor="F8CBAD"),
            "high": PatternFill("solid", fgColor="FCE4D6"),
            "medium": PatternFill("solid", fgColor="FFF2CC"),
        }

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        long_text_headers = {"source_path", "context", "error_msg", "rule_description", "scanned_details", "value"}
        sample_rows = rows[:1001]
        for col_idx, header_name in enumerate(header, start=1):
            header_name = str(header[col_idx - 1]) if col_idx <= len(header) else ""
            max_len = len(header_name)
            for row in sample_rows[1:]:
                value = "" if col_idx > len(row) or row[col_idx - 1] is None else str(row[col_idx - 1])
                max_len = max(max_len, min(len(value), 80))
            width = min(max(max_len + 2, 12), 60)
            if header_name in long_text_headers:
                width = min(max(width, 36), 70)
            worksheet.column_dimensions[self._column_name(col_idx)].width = width

        if "risk_level" in header and worksheet.max_row >= 2:
            risk_col = header.index("risk_level") + 1
            risk_col_letter = self._column_name(risk_col)
            data_range = f"A2:{self._column_name(worksheet.max_column)}{worksheet.max_row}"
            for risk_value, fill in risk_fills.items():
                worksheet.conditional_formatting.add(
                    data_range,
                    FormulaRule(
                        formula=[f'${risk_col_letter}2="{risk_value}"'],
                        fill=fill,
                    ),
                )
